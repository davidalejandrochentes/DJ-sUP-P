from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from .models import Venta, DetalleVenta, Cliente
from stock.models import Producto
from .forms import VentaForm, DetalleVentaForm
from django.db.models import Q
from django.core.exceptions import ValidationError
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill, Border, Side
from datetime import datetime
from django.views.decorators.cache import never_cache

# Create your views here.
@login_required
@never_cache
def venta(request):
    search_date_specific = request.GET.get('search_date_specific', '')
    search_date_start = request.GET.get('search_date_start', '')
    search_date_end = request.GET.get('search_date_end', '')
    search_code = request.GET.get('search_code', '')
    
    ventas_list = Venta.objects.filter(user=request.user)
    
    # Aplicar filtros de fecha
    if search_date_specific:
        ventas_list = ventas_list.filter(fecha__date=search_date_specific)
    else:
        if search_date_start:
            ventas_list = ventas_list.filter(fecha__date__gte=search_date_start)
        if search_date_end:
            ventas_list = ventas_list.filter(fecha__date__lte=search_date_end)
        
    # Aplicar filtros de código o cliente independientemente de las fechas
    if search_code:
        # Si la búsqueda contiene la palabra "minorista", incluir ventas sin cliente
        if 'desco' in search_code.lower():
            ventas_list = ventas_list.filter(
                Q(código__icontains=search_code) |
                Q(cliente__nombre__icontains=search_code) |
                Q(cliente__isnull=True)
            )
        else:
            ventas_list = ventas_list.filter(
                Q(código__icontains=search_code) |
                Q(cliente__nombre__icontains=search_code)
            )
        
    ventas_list = ventas_list.select_related('user', 'cliente').order_by('-fecha')

    page = request.GET.get('page', 1)
    paginator = Paginator(ventas_list, 20)
    try:
        ventas = paginator.page(page)
    except PageNotAnInteger:
        ventas = paginator.page(1)
    except EmptyPage:
        ventas = paginator.page(paginator.num_pages)

    context = {
        'ventas': ventas,
        'total': ventas_list.count(),
        'search_date_specific': search_date_specific,
        'search_date_start': search_date_start,
        'search_date_end': search_date_end,
        'search_code': search_code
    }
    return render(request, 'venta/venta.html', context)

@login_required
@never_cache
def nueva_venta(request):
    if request.method == 'POST':
        form = VentaForm(request.POST, user=request.user)
        if form.is_valid():
            venta = form.save(commit=False)
            venta.user = request.user
            
            # Si es cliente desconocido, asegurarse de que no se asigne cliente
            if form.cleaned_data['tipo_cliente'] == 'minorista':
                venta.cliente = None
            
            # Establecer la fecha actual
            venta.fecha = datetime.now()
                
            venta.save()
            return redirect('detalle_venta', venta_id=venta.id)
    else:
        form = VentaForm(user=request.user)
    
    return render(request, 'venta/venta_form.html', {'form': form, 'titulo': 'Nueva Venta'})

@login_required
@never_cache
def detalle_venta(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id, user=request.user)
    detalles = venta.detalles.all()
    
    # Initialize forms at the beginning
    form_detalle = DetalleVentaForm()
    form_detalle.fields['producto'].queryset = Producto.objects.filter(user=request.user)
    
    if request.method == 'POST':
        if 'editar_venta' in request.POST:
            form_venta = VentaForm(request.POST, instance=venta, user=request.user)
            if form_venta.is_valid():
                venta_editada = form_venta.save(commit=False)
                
                # Procesar fecha y hora
                try:
                    fecha = request.POST.get('fecha')
                    hora = request.POST.get('hora')
                    if fecha and hora:
                        fecha_hora = datetime.strptime(f"{fecha} {hora}", "%Y-%m-%d %H:%M")
                        venta_editada.fecha = fecha_hora
                    venta_editada.save()
                    messages.success(request, 'Venta actualizada correctamente.')
                except ValueError:
                    messages.error(request, 'Formato de fecha u hora inválido.')
                
                return redirect('detalle_venta', venta_id=venta.id)
        elif 'agregar_detalle' in request.POST:
            form_detalle = DetalleVentaForm(request.POST)
            form_detalle.fields['producto'].queryset = Producto.objects.filter(user=request.user)
            if form_detalle.is_valid():
                detalle = form_detalle.save(commit=False)
                detalle.venta = venta
                if detalle.cantidad > 0:
                    detalle.save()
                else:
                    messages.error(request, 'La cantidad debe ser mayor que 0.')
            else:
                for field, errors in form_detalle.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
            return redirect('detalle_venta', venta_id=venta.id)
    else:
        form_venta = VentaForm(instance=venta, user=request.user)

    context = {
        'venta': venta,
        'detalles': detalles,
        'form_venta': form_venta,
        'form_detalle': form_detalle,
        'titulo': f'Venta #{venta.código}'
    }
    return render(request, 'venta/detalle_venta.html', context)

@login_required
@never_cache
def eliminar_venta(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id, user=request.user)
    venta.delete()
    return redirect('venta')

@login_required
@never_cache
def eliminar_detalle(request, detalle_id):
    detalle = get_object_or_404(DetalleVenta, id=detalle_id, venta__user=request.user)
    venta_id = detalle.venta.id
    detalle.delete()
    return redirect('detalle_venta', venta_id=venta_id)

@login_required
@never_cache
def descargar_venta_excel(request, venta_id):
    venta = get_object_or_404(Venta, id=venta_id, user=request.user)
    
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Venta #{venta.código}"
    
    # Configurar estilos
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    table_header_style = NamedStyle(name='table_header_style')
    table_header_style.font = Font(bold=True)
    table_header_style.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    table_header_style.alignment = Alignment(horizontal='center', vertical='center')
    table_header_style.border = thin_border

    column_header_style = NamedStyle(name='column_header_style')
    column_header_style.font = Font(bold=True)
    column_header_style.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    column_header_style.alignment = Alignment(horizontal='center', vertical='center')
    column_header_style.border = thin_border

    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='left', vertical='center')
    data_style.border = thin_border
    
    # Escribir la información de la venta
    # Título de la primera tabla
    ws.merge_cells('A1:B1')
    ws['A1'] = "Información de la Venta"
    ws['A1'].style = table_header_style
    
    # Encabezados y datos de la primera tabla
    headers = [('A2', 'Código'), ('A3', 'Fecha'), ('A4', 'Cliente'), ('A5', 'Total')]
    for cell_ref, header in headers:
        ws[cell_ref] = header
        ws[cell_ref].style = column_header_style
    
    # Datos
    ws['B2'] = venta.código
    ws['B3'] = venta.fecha.strftime('%d/%m/%Y %H:%M')
    ws['B4'] = venta.cliente.nombre if venta.cliente else "Cliente Desconocido"
    ws['B5'] = f"${venta.total}"
    
    for row in range(2, 6):
        ws[f'B{row}'].style = data_style
    
    # Agregar dos filas en blanco
    current_row = 8
    
    # Título de la segunda tabla
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "Detalles de la Venta"
    ws[f'A{current_row}'].style = table_header_style
    
    # Encabezados de la segunda tabla
    current_row += 1
    headers = ['Producto', 'Cantidad', 'Precio Unitario', 'Subtotal']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.style = column_header_style
    
    # Agregar los detalles de la venta
    current_row += 1
    for detalle in venta.detalles.all():
        nombre_producto = detalle.producto.nombre if detalle.producto else f"{detalle.nombre_producto} (Producto eliminado)"
        ws.cell(row=current_row, column=1, value=nombre_producto).style = data_style
        ws.cell(row=current_row, column=2, value=detalle.cantidad).style = data_style
        ws.cell(row=current_row, column=3, value=f"${detalle.precio_unitario}").style = data_style
        ws.cell(row=current_row, column=4, value=f"${detalle.subtotal}").style = data_style
        current_row += 1
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Venta_{venta.código}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response

@login_required
@never_cache
def descargar_ventas_excel(request):
    search_date_specific = request.GET.get('search_date_specific', '')
    search_date_start = request.GET.get('search_date_start', '')
    search_date_end = request.GET.get('search_date_end', '')
    search_code = request.GET.get('search_code', '')
    
    ventas_list = Venta.objects.filter(user=request.user)
    
    # Aplicar filtros de fecha
    if search_date_specific:
        ventas_list = ventas_list.filter(fecha__date=search_date_specific)
    else:
        if search_date_start:
            ventas_list = ventas_list.filter(fecha__date__gte=search_date_start)
        if search_date_end:
            ventas_list = ventas_list.filter(fecha__date__lte=search_date_end)
    
    # Aplicar filtros de código o cliente
    if search_code:
        # Si la búsqueda contiene la palabra "minorista", incluir ventas sin cliente
        if 'desco' in search_code.lower():
            ventas_list = ventas_list.filter(
                Q(código__icontains=search_code) |
                Q(cliente__nombre__icontains=search_code) |
                Q(cliente__isnull=True)
            )
        else:
            ventas_list = ventas_list.filter(
                Q(código__icontains=search_code) |
                Q(cliente__nombre__icontains=search_code)
            )
    
    ventas_list = ventas_list.select_related('user', 'cliente').prefetch_related('detalles', 'detalles__producto').order_by('-fecha')
    
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Configurar estilos
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    header_style = NamedStyle(name='header_style')
    header_style.font = Font(bold=True)
    header_style.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.border = thin_border

    column_header_style = NamedStyle(name='column_header_style')
    column_header_style.font = Font(bold=True)
    column_header_style.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    column_header_style.alignment = Alignment(horizontal='center', vertical='center')
    column_header_style.border = thin_border

    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='left', vertical='top')
    data_style.border = thin_border

    total_style = NamedStyle(name='total_style')
    total_style.font = Font(bold=True)
    total_style.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    total_style.alignment = Alignment(horizontal='center', vertical='center')
    total_style.border = thin_border

    # Escribir los encabezados
    headers = ['Código', 'Fecha', 'Cliente', 'Total', 'Productos', 'Detalles']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.style = column_header_style

    # Configurar el ancho máximo para la columna de Productos y Detalles
    ws.column_dimensions['E'].width = 40  # Columna 'Productos'
    ws.column_dimensions['F'].width = 40  # Columna 'Detalles'

    # Agregar los datos de las ventas
    total_ventas = 0
    for row, venta in enumerate(ventas_list, 2):
        # Obtener los productos y sus cantidades
        productos = []
        detalles_precios = []
        for detalle in venta.detalles.all():
            nombre_producto = detalle.producto.nombre if detalle.producto else f"{detalle.nombre_producto} (Producto eliminado)"
            productos.append(f"{nombre_producto} (x{detalle.cantidad})")
            detalles_precios.append(f"${detalle.subtotal}")

        # Configurar las celdas
        for col in range(1, 7):
            ws.cell(row=row, column=col).style = data_style

        ws.cell(row=row, column=1, value=venta.código)
        ws.cell(row=row, column=2, value=venta.fecha.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=3, value=venta.cliente.nombre if venta.cliente else "Cliente Desconocido")
        ws.cell(row=row, column=4, value=f"${venta.total}")
        
        # Configurar wrap_text para Productos y Detalles
        productos_cell = ws.cell(row=row, column=5, value=", ".join(productos))
        productos_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        productos_cell.border = thin_border
        
        detalles_cell = ws.cell(row=row, column=6, value=", ".join(detalles_precios))
        detalles_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        detalles_cell.border = thin_border
        
        total_ventas += venta.total

    # Agregar una línea en blanco
    last_row = len(ventas_list) + 3

    # Agregar el total general con estilo
    total_label = ws.cell(row=last_row, column=3, value="TOTAL GENERAL:")
    total_label.style = total_style
    
    total_value = ws.cell(row=last_row, column=4, value=f"${total_ventas}")
    total_value.style = total_style
    
    # Ajustar ancho de columnas (excepto la de Productos y Detalles que ya tienen ancho fijo)
    for column in ws.columns:
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        if column_letter not in ['E', 'F']:  # Saltamos las columnas de Productos y Detalles
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Ventas_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response