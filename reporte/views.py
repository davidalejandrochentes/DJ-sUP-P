from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Sum, F, DecimalField, ExpressionWrapper, Case, When, Value, CharField
from django.db.models.functions import Concat
import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.views.decorators.cache import never_cache

from venta.models import DetalleVenta, Venta
from stock.models import Producto

#------------------------------------------------------------------------------

def reporte(request):
    return render(request, 'reporte/reporte.html', {})

#------------------------------------------------------------------------------

@login_required
@never_cache
def ganancias_por_cliente(request):
    periodo = request.GET.get('periodo', 'historico')
    
    fecha_inicio = None
    periodo_texto = "Histórico"
    
    if periodo != 'historico':
        hoy = timezone.now()
        if periodo == 'semana':
            fecha_inicio = hoy - timedelta(days=7)
            periodo_texto = "Última Semana"
        elif periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
            periodo_texto = "Último Mes"
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
            periodo_texto = "Último Trimestre"
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
            periodo_texto = "Último Semestre"
        elif periodo == 'anio':
            fecha_inicio = hoy - timedelta(days=365)
            periodo_texto = "Último Año"
    
    query = Venta.objects.filter(user=request.user)
    
    if fecha_inicio:
        query = query.filter(fecha__gte=fecha_inicio)
    
    clientes_ganancias = query.values(
        'cliente__nombre'
    ).annotate(
        nombre_completo=Case(
            When(cliente__nombre__isnull=True, then=Value('Clientes Desconocidos')),
            default=F('cliente__nombre'),
            output_field=CharField(),
        ),
        total_ganancias=Sum(F('total'))
    ).order_by('-total_ganancias')
    
    paginator = Paginator(clientes_ganancias, 20)
    page = request.GET.get('page', 1)
    
    try:
        clientes_paginados = paginator.page(page)
    except PageNotAnInteger:
        clientes_paginados = paginator.page(1)
    except EmptyPage:
        clientes_paginados = paginator.page(paginator.num_pages)
    
    clientes_top10 = list(clientes_ganancias[:10])
    labels = [item['nombre_completo'] for item in clientes_top10]
    data = [float(item['total_ganancias']) for item in clientes_top10]
    
    context = {
        'labels': labels,
        'data': data,
        'clientes': clientes_paginados,
        'total_clientes': clientes_ganancias.count(),
        'periodo': periodo,
        'periodo_texto': periodo_texto,
    }
    
    return render(request, 'reporte/ganancias_por_cliente.html', context)

#------------------------------------------------------------------------------    

@login_required
@never_cache
def productos_con_mayor_interes(request):
    productos = Producto.objects.filter(user=request.user).annotate(
        margen_ganancia=ExpressionWrapper(
            F('precio_de_venta') - F('precio_de_adquisición'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        ),
        porcentaje_ganancia=ExpressionWrapper(
            (F('precio_de_venta') - F('precio_de_adquisición')) * 100.0 / F('precio_de_adquisición'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    ).order_by('-margen_ganancia')
    
    paginator = Paginator(productos, 20)
    page = request.GET.get('page', 1)
    
    try:
        productos_paginados = paginator.page(page)
    except PageNotAnInteger:
        productos_paginados = paginator.page(1)
    except EmptyPage:
        productos_paginados = paginator.page(paginator.num_pages)
    
    productos_top10 = list(productos[:10])
    labels = [producto.nombre for producto in productos_top10]
    data = [float(producto.margen_ganancia) for producto in productos_top10]
    
    context = {
        'labels': labels,
        'data': data,
        'productos': productos_paginados,
        'total_productos': productos.count(),
    }
    
    return render(request, 'reporte/productos_con_mayor_interes.html', context)

#------------------------------------------------------------------------------

@login_required
@never_cache
def ingresos_por_producto(request):
    
    periodo = request.GET.get('periodo', 'historico')
    
    fecha_inicio = None
    periodo_texto = "Histórico"
    
    if periodo != 'historico':
        hoy = timezone.now()
        if periodo == 'semana':
            fecha_inicio = hoy - timedelta(days=7)
            periodo_texto = "Última Semana"
        elif periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
            periodo_texto = "Último Mes"
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
            periodo_texto = "Último Trimestre"
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
            periodo_texto = "Último Semestre"
        elif periodo == 'anio':
            fecha_inicio = hoy - timedelta(days=365)
            periodo_texto = "Último Año"
    
    query = DetalleVenta.objects.filter(venta__user=request.user)
    
    if fecha_inicio:
        query = query.filter(venta__fecha__gte=fecha_inicio)
    
    productos_ingresos = query.values(
        'producto__nombre',
        'producto__código',
        'producto__categoría__nombre',
        'nombre_producto',
        'producto'
    ).annotate(
        total_ingresos=Sum(F('cantidad') * F('precio_unitario')),
        nombre_mostrado=Case(
            When(producto__isnull=True, then=Concat('nombre_producto', Value(' (Producto eliminado)'))),
            default='producto__nombre',
            output_field=CharField(),
        )
    ).order_by('-total_ingresos')
    
    paginator = Paginator(productos_ingresos, 20)
    page = request.GET.get('page', 1)
    
    try:
        productos_paginados = paginator.page(page)
    except PageNotAnInteger:
        productos_paginados = paginator.page(1)
    except EmptyPage:
        productos_paginados = paginator.page(paginator.num_pages)
    
    productos_top10 = list(productos_ingresos[:10])
    labels = [item['nombre_mostrado'] for item in productos_top10]
    data = [float(item['total_ingresos']) for item in productos_top10]
    
    context = {
        'labels': labels,
        'data': data,
        'productos': productos_paginados,
        'total_productos': productos_ingresos.count(),
        'periodo': periodo,
        'periodo_texto': periodo_texto,
    }
    
    return render(request, 'reporte/ingresos_por_producto.html', context)

#------------------------------------------------------------------------------

@login_required
@never_cache
def productos_mas_vendidos(request):
    periodo = request.GET.get('periodo', 'historico')
    
    fecha_inicio = None
    periodo_texto = "Histórico"
    
    if periodo != 'historico':
        hoy = timezone.now()
        if periodo == 'semana':
            fecha_inicio = hoy - timedelta(days=7)
            periodo_texto = "Última Semana"
        elif periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
            periodo_texto = "Último Mes"
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
            periodo_texto = "Último Trimestre"
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
            periodo_texto = "Último Semestre"
        elif periodo == 'anio':
            fecha_inicio = hoy - timedelta(days=365)
            periodo_texto = "Último Año"
    
    query = DetalleVenta.objects.filter(venta__user=request.user)
    
    if fecha_inicio:
        query = query.filter(venta__fecha__gte=fecha_inicio)
    
    productos_vendidos = query.values(
        'producto__nombre',
        'producto__código',
        'producto__categoría__nombre',
        'nombre_producto',
        'producto'
    ).annotate(
        cantidad_vendida=Sum('cantidad'),
        nombre_mostrado=Case(
            When(producto__isnull=True, then=Concat('nombre_producto', Value(' (Producto eliminado)'))),
            default='producto__nombre',
            output_field=CharField(),
        )
    ).order_by('-cantidad_vendida')
    
    paginator = Paginator(productos_vendidos, 20)
    page = request.GET.get('page', 1)
    
    try:
        productos_paginados = paginator.page(page)
    except PageNotAnInteger:
        productos_paginados = paginator.page(1)
    except EmptyPage:
        productos_paginados = paginator.page(paginator.num_pages)
    
    productos_top10 = list(productos_vendidos[:10])
    labels = [item['nombre_mostrado'] for item in productos_top10]
    data = [float(item['cantidad_vendida']) for item in productos_top10]
    
    context = {
        'labels': labels,
        'data': data,
        'productos': productos_paginados,
        'total_productos': productos_vendidos.count(),
        'periodo': periodo,
        'periodo_texto': periodo_texto,
    }
    
    return render(request, 'reporte/productos_mas_vendidos.html', context)

#------------------------------------------------------------------------------

@login_required
@never_cache
def exportar_productos_vendidos_excel(request):
    periodo = request.GET.get('periodo', 'historico')
    
    fecha_inicio = None
    periodo_texto = "Histórico"
    
    if periodo != 'historico':
        hoy = timezone.now()
        if periodo == 'semana':
            fecha_inicio = hoy - timedelta(days=7)
            periodo_texto = "Última Semana"
        elif periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
            periodo_texto = "Último Mes"
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
            periodo_texto = "Último Trimestre"
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
            periodo_texto = "Último Semestre"
        elif periodo == 'anio':
            fecha_inicio = hoy - timedelta(days=365)
            periodo_texto = "Último Año"

    query = DetalleVenta.objects.filter(venta__user=request.user)
    
    if fecha_inicio:
        query = query.filter(venta__fecha__gte=fecha_inicio)
    
    productos_vendidos = query.values(
        'producto__nombre',
        'producto__código',
        'producto__categoría__nombre',
        'nombre_producto',
        'producto'
    ).annotate(
        cantidad_vendida=Sum('cantidad'),
        nombre_mostrado=Case(
            When(producto__isnull=True, then=Concat('nombre_producto', Value(' (Producto eliminado)'))),
            default='producto__nombre',
            output_field=CharField(),
        )
    ).order_by('-cantidad_vendida')
    
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = f"Productos Vendidos - {periodo_texto}"

    # Configurar estilos
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    table_header_style = NamedStyle(name='table_header_style')
    table_header_style.font = Font(bold=True)
    table_header_style.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    table_header_style.alignment = Alignment(horizontal='center', vertical='center')
    table_header_style.border = thin_border

    column_header_style = NamedStyle(name='column_header_style')
    column_header_style.font = Font(bold=True)
    column_header_style.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    column_header_style.alignment = Alignment(horizontal='center', vertical='center')
    column_header_style.border = thin_border

    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='left', vertical='center')
    data_style.border = thin_border

    number_style = NamedStyle(name='number_style')
    number_style.alignment = Alignment(horizontal='right', vertical='center')
    number_style.border = thin_border

    total_style = NamedStyle(name='total_style')
    total_style.font = Font(bold=True)
    total_style.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    total_style.alignment = Alignment(horizontal='center', vertical='center')
    total_style.border = thin_border

    # Título del reporte
    ws.merge_cells('A1:B1')
    ws['A1'] = f"Reporte de Productos Vendidos - {periodo_texto}"
    ws['A1'].style = table_header_style

    # Encabezados de columnas
    headers = ['Producto', 'Cantidad Vendida', 'Categoría', 'Código de Producto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.style = column_header_style

    # Agregar los datos
    row = 3
    total_vendido = 0
    for producto in productos_vendidos:
        # Nombre del producto
        ws.cell(row=row, column=1, value=producto['nombre_mostrado']).style = data_style
        
        # Cantidad vendida
        cantidad = producto['cantidad_vendida']
        ws.cell(row=row, column=2, value=cantidad).style = number_style
        
        # Categoría
        cell_categoria = ws.cell(row=row, column=3, value=producto['producto__categoría__nombre'] or '-')
        cell_categoria.style = data_style
        cell_categoria.alignment = Alignment(horizontal='right')
        
        # Código del producto
        cell_codigo = ws.cell(row=row, column=4, value=producto['producto__código'] or '-')
        cell_codigo.style = data_style
        cell_codigo.alignment = Alignment(horizontal='right')
        
        total_vendido += cantidad
        row += 1

    # Agregar el total general
    ws.merge_cells(f'A{row+1}:A{row+1}')
    total_label = ws.cell(row=row+1, column=1, value="TOTAL GENERAL")
    total_label.style = total_style
    
    total_value = ws.cell(row=row+1, column=2, value=total_vendido)
    total_value.style = total_style

    # Ajustar ancho de columnas basado en el contenido y los títulos
    headers = ['Producto', 'Cantidad Vendida', 'Categoría', 'Código de Producto']
    for col, header in enumerate(headers, 1):
        column_letter = get_column_letter(col)
        # Establecer un ancho mínimo basado en la longitud del encabezado
        header_length = len(header)
        min_width = header_length + 2  # Agregamos 2 para un poco de padding
        
        # Revisar el contenido de la columna para asegurar que todo quepa
        max_length = min_width
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Usar el valor más grande entre el ancho mínimo y el contenido más largo
        ws.column_dimensions[column_letter].width = max(min_width, max_length)
        
    # Ajustar el ancho de la columna 'Categoría' para que sea un poco más grande
    ws.column_dimensions[get_column_letter(3)].width += 15;

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=productos_vendidos_{periodo}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

#------------------------------------------------------------------------------

@login_required
@never_cache
def exportar_ingresos_producto_excel(request):
    periodo = request.GET.get('periodo', 'historico')
    
    fecha_inicio = None
    periodo_texto = "Histórico"
    
    if periodo != 'historico':
        hoy = timezone.now()
        if periodo == 'semana':
            fecha_inicio = hoy - timedelta(days=7)
            periodo_texto = "Última Semana"
        elif periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
            periodo_texto = "Último Mes"
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
            periodo_texto = "Último Trimestre"
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
            periodo_texto = "Último Semestre"
        elif periodo == 'anio':
            fecha_inicio = hoy - timedelta(days=365)
            periodo_texto = "Último Año"

    query = DetalleVenta.objects.filter(venta__user=request.user)
    
    if fecha_inicio:
        query = query.filter(venta__fecha__gte=fecha_inicio)
    
    ingresos_productos = query.values(
        'producto__nombre',
        'producto__código',
        'producto__categoría__nombre',
        'nombre_producto',
        'producto'
    ).annotate(
        total_ingresos=Sum(F('cantidad') * F('precio_unitario')),
        nombre_mostrado=Case(
            When(producto__isnull=True, then=Concat(F('nombre_producto'), Value(' (Producto eliminado)'))),
            default=F('producto__nombre'),
            output_field=CharField(),
        )
    ).order_by('-total_ingresos')
    
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ingresos por Producto - {periodo_texto}"

    # Configurar estilos
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    table_header_style = NamedStyle(name='table_header_style')
    table_header_style.font = Font(bold=True)
    table_header_style.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    table_header_style.alignment = Alignment(horizontal='center', vertical='center')
    table_header_style.border = thin_border

    column_header_style = NamedStyle(name='column_header_style')
    column_header_style.font = Font(bold=True)
    column_header_style.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    column_header_style.alignment = Alignment(horizontal='center', vertical='center')
    column_header_style.border = thin_border

    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='left', vertical='center')
    data_style.border = thin_border

    money_style = NamedStyle(name='money_style')
    money_style.alignment = Alignment(horizontal='right', vertical='center')
    money_style.border = thin_border

    total_style = NamedStyle(name='total_style')
    total_style.font = Font(bold=True)
    total_style.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    total_style.alignment = Alignment(horizontal='center', vertical='center')
    total_style.border = thin_border

    # Título del reporte
    ws.merge_cells('A1:B1')
    ws['A1'] = f"Reporte de Ingresos por Producto - {periodo_texto}"
    ws['A1'].style = table_header_style

    # Encabezados de columnas
    headers = ['Producto', 'Ingresos ($)', 'Categoría', 'Código de Producto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.style = column_header_style

    # Agregar los datos
    row = 3
    total_ingresos = 0
    for producto in ingresos_productos:
        # Nombre del producto
        ws.cell(row=row, column=1, value=producto['nombre_mostrado']).style = data_style
        
        # Ingresos
        ingresos = float(producto['total_ingresos'])
        ws.cell(row=row, column=2, value=f"${ingresos:,.2f}").style = money_style
        
        # Categoría
        cell_categoria = ws.cell(row=row, column=3, value=producto['producto__categoría__nombre'] or '-')
        cell_categoria.style = data_style
        cell_categoria.alignment = Alignment(horizontal='right')
        
        # Código del producto
        cell_codigo = ws.cell(row=row, column=4, value=producto['producto__código'] or '-')
        cell_codigo.style = data_style
        cell_codigo.alignment = Alignment(horizontal='right')
        
        total_ingresos += ingresos
        row += 1

    # Agregar el total general
    ws.merge_cells(f'A{row+1}:A{row+1}')
    total_label = ws.cell(row=row+1, column=1, value="TOTAL GENERAL")
    total_label.style = total_style
    
    total_value = ws.cell(row=row+1, column=2, value=f"${total_ingresos:,.2f}")
    total_value.style = total_style

    # Ajustar ancho de columnas basado en el contenido y los títulos
    headers = ['Producto', 'Ingresos ($)', 'Categoría', 'Código de Producto']
    for col, header in enumerate(headers, 1):
        column_letter = get_column_letter(col)
        # Establecer un ancho mínimo basado en la longitud del encabezado
        header_length = len(header)
        min_width = header_length + 2  # Agregamos 2 para un poco de padding
        
        # Revisar el contenido de la columna para asegurar que todo quepa
        max_length = min_width
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Usar el valor más grande entre el ancho mínimo y el contenido más largo
        ws.column_dimensions[column_letter].width = max(min_width, max_length)

    # Ajustar el ancho de la columna 'Categoría' para que sea un poco más grande
    ws.column_dimensions[get_column_letter(3)].width += 15;

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ingresos_productos_{periodo}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

#------------------------------------------------------------------------------

@login_required
@never_cache
def exportar_productos_margen_excel(request):
    # Obtener los productos con sus márgenes
    productos = Producto.objects.filter(user=request.user).annotate(
        margen_ganancia=ExpressionWrapper(
            F('precio_de_venta') - F('precio_de_adquisición'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        ),
        porcentaje_ganancia=ExpressionWrapper(
            (F('precio_de_venta') - F('precio_de_adquisición')) * 100.0 / F('precio_de_adquisición'),
            output_field=DecimalField(max_digits=10, decimal_places=2)
        )
    ).order_by('-margen_ganancia')
    
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Configurar estilos
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    table_header_style = NamedStyle(name='table_header_style')
    table_header_style.font = Font(bold=True)
    table_header_style.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    table_header_style.alignment = Alignment(horizontal='center', vertical='center')
    table_header_style.border = thin_border

    column_header_style = NamedStyle(name='column_header_style')
    column_header_style.font = Font(bold=True)
    column_header_style.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    column_header_style.alignment = Alignment(horizontal='center', vertical='center')
    column_header_style.border = thin_border

    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='left', vertical='center')
    data_style.border = thin_border

    money_style = NamedStyle(name='money_style')
    money_style.alignment = Alignment(horizontal='right', vertical='center')
    money_style.border = thin_border

    percentage_style = NamedStyle(name='percentage_style')
    percentage_style.alignment = Alignment(horizontal='right', vertical='center')
    percentage_style.border = thin_border

    total_style = NamedStyle(name='total_style')
    total_style.font = Font(bold=True)
    total_style.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    total_style.alignment = Alignment(horizontal='center', vertical='center')
    total_style.border = thin_border

    # Título del reporte
    ws.merge_cells('A1:G1')
    ws['A1'] = "Reporte de Productos por Margen de Ganancia"
    ws['A1'].style = table_header_style

    # Encabezados de columnas
    headers = ['Producto', 'Precio de Adquisición', 'Precio de Venta', 'Margen de Ganancia ($)', 'Margen de Ganancia (%)', 'Código de Producto', 'Categoría']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.style = column_header_style

    # Agregar los datos
    row = 3
    for producto in productos:
        # Nombre del producto
        ws.cell(row=row, column=1, value=producto.nombre).style = data_style
        
        # Precios y márgenes
        ws.cell(row=row, column=2, value=f"${float(producto.precio_de_adquisición):,.2f}").style = money_style
        ws.cell(row=row, column=3, value=f"${float(producto.precio_de_venta):,.2f}").style = money_style
        ws.cell(row=row, column=4, value=f"${float(producto.margen_ganancia):,.2f}").style = money_style
        ws.cell(row=row, column=5, value=f"{float(producto.porcentaje_ganancia):.1f}%").style = percentage_style
        
        # Código del producto
        cell_codigo = ws.cell(row=row, column=6, value=producto.código)
        cell_codigo.style = data_style
        cell_codigo.alignment = Alignment(horizontal='right')
        
        # Categoría
        cell_categoria = ws.cell(row=row, column=7, value=producto.categoría.nombre)
        cell_categoria.style = data_style
        cell_categoria.alignment = Alignment(horizontal='right')
        
        row += 1

    # Ajustar ancho de columnas basado en el contenido y los títulos
    headers = ['Producto', 'Precio de Adquisición', 'Precio de Venta', 'Margen de Ganancia ($)', 'Margen de Ganancia (%)', 'Código de Producto', 'Categoría']
    for col, header in enumerate(headers, 1):
        column_letter = get_column_letter(col)
        # Establecer un ancho mínimo basado en la longitud del encabezado
        header_length = len(header)
        min_width = header_length + 2  # Agregamos 2 para un poco de padding
        
        # Revisar el contenido de la columna para asegurar que todo quepa
        max_length = min_width
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Usar el valor más grande entre el ancho mínimo y el contenido más largo
        ws.column_dimensions[column_letter].width = max(min_width, max_length)

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=productos_margen_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

#------------------------------------------------------------------------------

@login_required
@never_cache
def exportar_ganancias_cliente_excel(request):
    periodo = request.GET.get('periodo', 'historico')
    
    fecha_inicio = None
    periodo_texto = "Histórico"
    
    if periodo != 'historico':
        hoy = timezone.now()
        if periodo == 'semana':
            fecha_inicio = hoy - timedelta(days=7)
            periodo_texto = "Última Semana"
        elif periodo == 'mes':
            fecha_inicio = hoy - timedelta(days=30)
            periodo_texto = "Último Mes"
        elif periodo == 'trimestre':
            fecha_inicio = hoy - timedelta(days=90)
            periodo_texto = "Último Trimestre"
        elif periodo == 'semestre':
            fecha_inicio = hoy - timedelta(days=180)
            periodo_texto = "Último Semestre"
        elif periodo == 'anio':
            fecha_inicio = hoy - timedelta(days=365)
            periodo_texto = "Último Año"
    
    query = Venta.objects.filter(user=request.user)
    
    if fecha_inicio:
        query = query.filter(fecha__gte=fecha_inicio)
    
    clientes_ganancias = query.values(
        'cliente__nombre'
    ).annotate(
        nombre_completo=Case(
            When(cliente__nombre__isnull=True, then=Value('Clientes Desconocidos')),
            default=F('cliente__nombre'),
            output_field=CharField(),
        ),
        total_ganancias=Sum(F('total'))
    ).order_by('-total_ganancias')
    
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ganancias - {periodo_texto}"
    
    # Configurar estilos
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    table_header_style = NamedStyle(name='table_header_style')
    table_header_style.font = Font(bold=True)
    table_header_style.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    table_header_style.alignment = Alignment(horizontal='center', vertical='center')
    table_header_style.border = thin_border

    column_header_style = NamedStyle(name='column_header_style')
    column_header_style.font = Font(bold=True)
    column_header_style.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    column_header_style.alignment = Alignment(horizontal='center', vertical='center')
    column_header_style.border = thin_border

    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='left', vertical='center')
    data_style.border = thin_border

    total_style = NamedStyle(name='total_style')
    total_style.font = Font(bold=True)
    total_style.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    total_style.alignment = Alignment(horizontal='center', vertical='center')
    total_style.border = thin_border

    money_style = NamedStyle(name='money_style')
    money_style.alignment = Alignment(horizontal='right', vertical='center')
    money_style.border = thin_border
    
    # Título del reporte
    ws.merge_cells('A1:B1')
    ws['A1'] = f"Reporte de Ganancias por Cliente - {periodo_texto}"
    ws['A1'].style = table_header_style

    # Encabezados de columnas
    headers = ['Cliente', 'Ganancias ($)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.style = column_header_style

    # Agregar los datos
    row = 3
    total_general = 0
    
    for cliente in clientes_ganancias:
        # Nombre del cliente
        ws.cell(row=row, column=1, value=cliente['nombre_completo']).style = data_style
        
        # Ganancias
        ganancia_cell = ws.cell(row=row, column=2, value=f"${float(cliente['total_ganancias']):,.2f}")
        ganancia_cell.style = money_style
        
        total_general += float(cliente['total_ganancias'])
        row += 1

    # Agregar el total general
    ws.merge_cells(f'A{row+1}:A{row+1}')
    total_label = ws.cell(row=row+1, column=1, value="TOTAL GENERAL")
    total_label.style = total_style
    
    total_value = ws.cell(row=row+1, column=2, value=f"${total_general:,.2f}")
    total_value.style = total_style

    # Ajustar ancho de columnas basado en el contenido y los títulos
    headers = ['Cliente', 'Ganancias ($)']
    for col, header in enumerate(headers, 1):
        column_letter = get_column_letter(col)
        # Establecer un ancho mínimo basado en la longitud del encabezado
        header_length = len(header)
        min_width = header_length + 2  # Agregamos 2 para un poco de padding
        
        # Revisar el contenido de la columna para asegurar que todo quepa
        max_length = min_width
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Usar el valor más grande entre el ancho mínimo y el contenido más largo
        ws.column_dimensions[column_letter].width = max(min_width, max_length)

    # Ajustar el ancho de la columna 'Categoría' para que sea un poco más grande
    ws.column_dimensions[get_column_letter(3)].width += 15;

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ganancias_clientes_{periodo}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response