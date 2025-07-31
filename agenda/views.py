from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Cliente, Proveedor, PalabraClave
from .forms import ClienteForm, ProveedorForm
from django.contrib import messages
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from django.http import HttpResponse
from datetime import datetime
from django.views.decorators.cache import never_cache

# Create your views here.
@login_required
@never_cache
def proveedores(request):
    search_query = request.GET.get('search', '')
    proveedores_list = Proveedor.objects.filter(
        user=request.user
    ).filter(
        Q(nombre__icontains=search_query) |
        Q(palabras_clave__palabra__icontains=search_query) |
        Q(teléfono__icontains=search_query)
    ).distinct().select_related('user').prefetch_related('palabras_clave').order_by('nombre')
    
    page = request.GET.get('page', 1)
    paginator = Paginator(proveedores_list, 20)
    try:
        proveedores = paginator.page(page)
    except PageNotAnInteger:
        proveedores = paginator.page(1)
    except EmptyPage:
        proveedores = paginator.page(paginator.num_pages)    
    context = {
        'proveedores': proveedores,
        'total': proveedores_list.count()
    }
    return render(request, 'agenda/proveedores.html', context)

@login_required
@never_cache
def proveedor(request, id):
    proveedor = get_object_or_404(Proveedor.objects.select_related('user'), id=id, user=request.user)
    context = {
        'proveedor': proveedor
    }
    return render(request, 'agenda/proveedor.html', context)

@login_required
@never_cache
def eliminar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor.objects.select_related('user'), id=id, user=request.user)
    proveedor.delete()
    return redirect ('proveedores') 

@login_required
@never_cache
def nuevo_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save(commit=False)
            proveedor.user = request.user
            proveedor.save()
            return redirect('proveedor', id=proveedor.id)
        else:
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    else:
        form = ProveedorForm()
    context = {'form': form}
    return render(request, 'agenda/nuevo_proveedor.html', context)

@login_required
@never_cache
def editar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor.objects.select_related('user'), id=id, user=request.user)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect('proveedor', id=proveedor.id)
        else:
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    else:
        form = ProveedorForm(instance=proveedor)
    context = {'form': form}
    return render(request, 'agenda/editar_proveedor.html', context)

@login_required
@never_cache
def agregar_palabra_clave_proveedor(request, proveedor_id):
    if request.method == 'POST':
        palabra = request.POST.get('palabra', '').strip().lower()
        if palabra:
            proveedor = get_object_or_404(Proveedor, id=proveedor_id, user=request.user)
            palabra_clave, created = PalabraClave.objects.get_or_create(
                user=request.user,
                palabra=palabra
            )
            proveedor.palabras_clave.add(palabra_clave)
    return redirect('proveedor', id=proveedor_id)

@login_required
@never_cache
def eliminar_palabra_clave_proveedor(request, proveedor_id, palabra_id):
    proveedor = get_object_or_404(Proveedor, id=proveedor_id, user=request.user)
    palabra_clave = get_object_or_404(PalabraClave, id=palabra_id, user=request.user)
    proveedor.palabras_clave.remove(palabra_clave)
    return redirect('proveedor', id=proveedor_id)

@login_required
@never_cache
def exportar_clientes_excel(request):
    # Obtener los clientes
    clientes = Cliente.objects.filter(user=request.user).order_by('nombre')
    
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Configurar estilos
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
    data_style = NamedStyle(name='data_style', 
                           border=thin_border,
                           alignment=Alignment(horizontal='left', vertical='top'))
    wrap_style = NamedStyle(name='wrap_style', 
                           border=thin_border,
                           alignment=Alignment(wrap_text=True, horizontal='left', vertical='top'))
    
    # Configurar ancho de columnas
    ws.column_dimensions['A'].width = 30  # Nombre
    ws.column_dimensions['B'].width = 15  # Teléfono
    ws.column_dimensions['C'].width = 30  # Email
    ws.column_dimensions['D'].width = 40  # Dirección
    ws.column_dimensions['E'].width = 40  # Notas
    ws.column_dimensions['F'].width = 30  # Palabras Clave
    
    # Agregar encabezados
    headers = ['Nombre', 'Teléfono', 'Email', 'Dirección', 'Notas', 'Palabras Clave']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Agregar datos
    row = 2
    for cliente in clientes:
        # Nombre
        ws.cell(row=row, column=1, value=cliente.nombre).style = data_style
        
        # Teléfono
        ws.cell(row=row, column=2, value=cliente.teléfono or '-').style = data_style
        
        # Email
        ws.cell(row=row, column=3, value=cliente.email or '-').style = data_style
        
        # Dirección
        ws.cell(row=row, column=4, value=cliente.dirección or '-').style = wrap_style
        
        # Notas
        ws.cell(row=row, column=5, value=cliente.notas or '-').style = wrap_style
        
        # Palabras clave
        palabras = ', '.join([p.palabra for p in cliente.palabras_clave.all()])
        ws.cell(row=row, column=6, value=palabras or '-').style = wrap_style
        
        row += 1
    
    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

@login_required
@never_cache
def exportar_proveedores_excel(request):
    # Obtener los proveedores
    proveedores = Proveedor.objects.filter(user=request.user).order_by('nombre')
    
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    # Configurar estilos
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
    data_style = NamedStyle(name='data_style', 
                           border=thin_border,
                           alignment=Alignment(horizontal='left', vertical='top'))
    wrap_style = NamedStyle(name='wrap_style', 
                           border=thin_border,
                           alignment=Alignment(wrap_text=True, horizontal='left', vertical='top'))
    
    # Configurar ancho de columnas
    ws.column_dimensions['A'].width = 30  # Nombre
    ws.column_dimensions['B'].width = 15  # Teléfono
    ws.column_dimensions['C'].width = 30  # Email
    ws.column_dimensions['D'].width = 40  # Dirección
    ws.column_dimensions['E'].width = 40  # Notas
    ws.column_dimensions['F'].width = 30  # Palabras Clave
    
    # Agregar encabezados
    headers = ['Nombre', 'Teléfono', 'Email', 'Dirección', 'Notas', 'Palabras Clave']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Agregar datos
    row = 2
    for proveedor in proveedores:
        # Nombre
        ws.cell(row=row, column=1, value=proveedor.nombre).style = data_style
        
        # Teléfono
        ws.cell(row=row, column=2, value=proveedor.teléfono or '-').style = data_style
        
        # Email
        ws.cell(row=row, column=3, value=proveedor.email or '-').style = data_style
        
        # Dirección
        ws.cell(row=row, column=4, value=proveedor.dirección or '-').style = wrap_style
        
        # Notas
        ws.cell(row=row, column=5, value=proveedor.notas or '-').style = wrap_style
        
        # Palabras clave
        palabras = ', '.join([p.palabra for p in proveedor.palabras_clave.all()])
        ws.cell(row=row, column=6, value=palabras or '-').style = wrap_style
        
        row += 1
    
    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Proveedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response

#--------------------------------------------------------------------------------------------------

@login_required
@never_cache
def clientes(request):
    search_query = request.GET.get('search', '')
    clientes_list = Cliente.objects.filter(
        user=request.user
    ).filter(
        Q(nombre__icontains=search_query) |
        Q(palabras_clave__palabra__icontains=search_query) |
        Q(teléfono__icontains=search_query)
    ).distinct().select_related('user').prefetch_related('palabras_clave').order_by('nombre')
    
    page = request.GET.get('page', 1)
    paginator = Paginator(clientes_list, 20)
    try:
        clientes = paginator.page(page)
    except PageNotAnInteger:
        clientes = paginator.page(1)
    except EmptyPage:
        clientes = paginator.page(paginator.num_pages)    
    context = {
        'clientes': clientes,
        'total': clientes_list.count()
    }
    return render(request, 'agenda/clientes.html', context)


@login_required
@never_cache
def cliente(request, id):
    cliente = get_object_or_404(Cliente.objects.select_related('user'), id=id, user=request.user)
    context = {
        'cliente': cliente
    }
    return render(request, 'agenda/cliente.html', context)

@login_required
@never_cache
def eliminar_cliente(request, id):
    cliente = get_object_or_404(Cliente.objects.select_related('user'), id=id, user=request.user)
    cliente.delete()
    return redirect ('clientes') 

@login_required
@never_cache
def nuevo_cliente(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save(commit=False)
            cliente.user = request.user
            cliente.save()
            return redirect('cliente', id=cliente.id)
        else:
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    else:
        form = ClienteForm()
    context = {'form': form}
    return render(request, 'agenda/nuevo_cliente.html', context)

@login_required
@never_cache
def editar_cliente(request, id):
    cliente = get_object_or_404(Cliente.objects.select_related('user'), id=id, user=request.user)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            return redirect('cliente', id=cliente.id)
        else:
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    else:
        form = ClienteForm(instance=cliente)
    context = {'form': form, 'cliente': cliente}
    return render(request, 'agenda/editar_cliente.html', context)

@login_required
@never_cache
def agregar_palabra_clave(request, cliente_id):
    if request.method == 'POST':
        palabra = request.POST.get('palabra', '').strip().lower()
        if palabra:
            cliente = get_object_or_404(Cliente, id=cliente_id, user=request.user)
            palabra_clave, created = PalabraClave.objects.get_or_create(
                user=request.user,
                palabra=palabra
            )
            cliente.palabras_clave.add(palabra_clave)
    return redirect('cliente', id=cliente_id)

@login_required
@never_cache
def eliminar_palabra_clave(request, cliente_id, palabra_id):
    cliente = get_object_or_404(Cliente, id=cliente_id, user=request.user)
    palabra_clave = get_object_or_404(PalabraClave, id=palabra_id, user=request.user)
    cliente.palabras_clave.remove(palabra_clave)
    return redirect('cliente', id=cliente_id)
