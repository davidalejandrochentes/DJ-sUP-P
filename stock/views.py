from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Categoria, Producto
from .forms import CategoriaForm, ProductoForm
from django.contrib import messages
from django.db.models import Count, F, Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.views.decorators.cache import never_cache

@login_required
@never_cache
def categorias_productos(request):
    categorias_list = Categoria.objects.filter(
        user=request.user, 
        nombre__icontains=request.GET.get('search', '')
    ).annotate(
        productos_bajo_stock=Count(
            'producto',
            filter=Q(producto__stock__lt=F('producto__alerta_stock')),
        )
    ).select_related('user').order_by('nombre')
    productos_bajo_stock = sum(c.productos_bajo_stock for c in categorias_list)
    
    page = request.GET.get('page', 1)
    paginator = Paginator(categorias_list, 20)
    try:
        categorias = paginator.page(page)
    except PageNotAnInteger:
        categorias = paginator.page(1)
    except EmptyPage:
        categorias = paginator.page(paginator.num_pages)
    context = {
        'categorias': categorias,
        'productos_bajo_stock': productos_bajo_stock,
        'total': categorias_list.count()
    }
    return render(request, 'stock/categorías.html', context)

@login_required
@never_cache
def eliminar_categoria(request, id):
    categoria = get_object_or_404(Categoria.objects.select_related('user'), id=id, user=request.user)
    categoria.delete()
    return redirect ('categorias_productos')

@login_required
@never_cache
def nueva_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            categoria = form.save(commit=False)
            categoria.user = request.user
            categoria.save()
            return redirect('categorias_productos')
        else:
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    else:
        form = CategoriaForm()
    context = {'form': form}
    return render(request, 'stock/nueva_categoria.html', context)

#--------------------------------------------------------------------------------------------------

@login_required
@never_cache
def lista_productos(request, id):
    categoria = get_object_or_404(Categoria.objects.select_related('user'), user=request.user, id=id)
    form = CategoriaForm(instance=categoria)
    productos_list = Producto.objects.filter(
        Q(user=request.user),
        Q(categoría=id),
        Q(nombre__icontains=request.GET.get('search', '')) | Q(código__icontains=request.GET.get('search', ''))
    ).select_related('user').order_by('nombre')
    
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            return redirect('lista_productos', id=categoria.id)
        else:
            return redirect(request.META.get('HTTP_REFERER', 'lista_productos'))
    alertas = [
        {
            'existencia': existencia,
            'existencia_fisica': existencia.stock,
        }
        for existencia in productos_list if existencia.stock <= existencia.alerta_stock
    ]

    page = request.GET.get('page', 1)
    paginator = Paginator(productos_list, 20)
    try:
        productos = paginator.page(page)
    except PageNotAnInteger:
        productos = paginator.page(1)
    except EmptyPage:
        productos = paginator.page(paginator.num_pages)

    context = {
        'productos': productos,
        'total': productos_list.count(),
        'categoria': categoria,
        'form': form,
        'total_alertas': len(alertas),
    }
    return render(request, "stock/productos.html", context)


@login_required
@never_cache
def nuevo_producto(request, id):
    categoria = get_object_or_404(Categoria.objects.select_related('user'), user=request.user, id=id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, user=request.user)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.user = request.user
            producto.categoría = categoria
            producto.save()
            return redirect('producto', id=producto.id)
        else:
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    else:
        form = ProductoForm(user=request.user)
    context = {
        'form': form,
        'categoria': categoria,
    }
    return render(request, 'stock/nuevo_producto.html', context)

#-------------------------------------------------------------------------------------          

@login_required 
@never_cache
def producto(request, id):
    producto = get_object_or_404(Producto.objects.select_related('user'), id=id, user=request.user)
    context = {
        'producto': producto
    }
    return render(request, 'stock/producto.html', context)

@login_required
@never_cache
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto.objects.select_related('user'), id=id, user=request.user)
    producto.delete()
    return redirect('lista_productos', id=producto.categoría.id)

@login_required
@never_cache
def productos_con_stock_bajo_general(request):
    productos_bajo_stock_list = Producto.objects.filter(user=request.user, stock__lt=F('alerta_stock')).select_related('user').order_by('nombre')
    page = request.GET.get('page', 1)
    paginator = Paginator(productos_bajo_stock_list, 20)
    try:
        productos_bajo_stock = paginator.page(page)
    except PageNotAnInteger:
        productos_bajo_stock = paginator.page(1)
    except EmptyPage:
        productos_bajo_stock = paginator.page(paginator.num_pages) 
    context = {
        'productos': productos_bajo_stock,
    }
    return render(request, 'stock/stock_bajo.html', context)

@login_required
@never_cache
def productos_con_stock_bajo(request, id):
    productos_bajo_stock_list = Producto.objects.filter(user=request.user, categoría=id, stock__lt=F('alerta_stock')).select_related('user').order_by('nombre')
    categoria = categoria = get_object_or_404(Categoria.objects.select_related('user'), id=id, user=request.user)
    page = request.GET.get('page', 1)
    paginator = Paginator(productos_bajo_stock_list, 20)
    try:
        productos_bajo_stock = paginator.page(page)
    except PageNotAnInteger:
        productos_bajo_stock = paginator.page(1)
    except EmptyPage:
        productos_bajo_stock = paginator.page(paginator.num_pages)
    context = {
        'productos': productos_bajo_stock,
        'categoria': categoria
    }
    return render(request, 'stock/productos_stock_bajo.html', context)

@login_required
@never_cache
def editar_producto(request, id):
    producto = get_object_or_404(Producto.objects.select_related('user'), id=id, user=request.user)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('producto', id=id)
        else:
            messages.error(request, "Alguno de los datos introducidos no son válidos, revise nuevamente cada campo")
    else:
        form = ProductoForm(instance=producto, user=request.user)
    context = {
        'form': form,
        'producto': producto,
    }
    return render(request, 'stock/editar_producto.html', context)

def get_proveedor_nombre(proveedor):
    return proveedor.nombre if proveedor else "Eliminado"

@login_required
@never_cache
def exportar_productos_excel(request):
    # Obtener todos los productos del usuario
    productos = Producto.objects.filter(user=request.user).select_related('categoría', 'proveedor')
    
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorías"
    
    # Definir estilos
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_style.alignment = Alignment(horizontal='left', vertical='center')
    header_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Estilo para las celdas de datos
    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='left', vertical='center')
    data_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definir encabezados
    headers = [
        'Código',
        'Nombre',
        'Categoría',
        'Proveedor',
        'Descripción',
        'Precio de Adquisición',
        'Precio de Venta',
        'Unidad de Medida',
        'Stock Actual',
        'Alerta de Stock'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = header_style
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Escribir datos
    for row, producto in enumerate(productos, 2):
        cells = [
            (producto.código),
            (producto.nombre),
            (producto.categoría.nombre),
            (get_proveedor_nombre(producto.proveedor)),
            (producto.descripción or ""),
            (float(producto.precio_de_adquisición)),
            (float(producto.precio_de_venta)),
            (producto.unidad_de_medida),
            (producto.stock),
            (producto.alerta_stock)
        ]
        
        for col, value in enumerate(cells, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.style = data_style
    
    # Ajustar el ancho de las columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Categorías.xlsx'
    
    # Guardar el libro de Excel
    wb.save(response)
    return response

@login_required
@never_cache
def exportar_productos_categoria_excel(request, id):
    # Obtener la categoría y verificar que pertenece al usuario
    categoria = get_object_or_404(Categoria, id=id, user=request.user)
    
    # Obtener los productos de la categoría
    productos = Producto.objects.filter(
        user=request.user,
        categoría=categoria
    ).select_related('categoría', 'proveedor')
    
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Productos - {categoria.nombre}"
    
    # Definir estilos
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_style.alignment = Alignment(horizontal='left', vertical='center')
    header_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Estilo para las celdas de datos
    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='left', vertical='center')
    data_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definir encabezados
    headers = [
        'Código',
        'Nombre',
        'Proveedor',
        'Descripción',
        'Precio de Adquisición',
        'Precio de Venta',
        'Unidad de Medida',
        'Stock Actual',
        'Alerta de Stock'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = header_style
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Escribir datos
    for row, producto in enumerate(productos, 2):
        cells = [
            (producto.código),
            (producto.nombre),
            (get_proveedor_nombre(producto.proveedor)),
            (producto.descripción or ""),
            (float(producto.precio_de_adquisición)),
            (float(producto.precio_de_venta)),
            (producto.unidad_de_medida),
            (producto.stock),
            (producto.alerta_stock)
        ]
        
        for col, value in enumerate(cells, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.style = data_style
    
    # Ajustar el ancho de las columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Productos_{categoria.nombre}.xlsx'
    
    # Guardar el libro de Excel
    wb.save(response)
    return response

@login_required
@never_cache
def exportar_productos_stock_bajo_excel(request):
    # Obtener productos con stock bajo
    productos = Producto.objects.filter(
        user=request.user,
        stock__lt=F('alerta_stock')
    ).select_related('categoría', 'proveedor').order_by('nombre')
    
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos Stock Bajo"
    
    # Definir estilos
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_style.alignment = Alignment(horizontal='left', vertical='center')
    header_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Estilo para las celdas de datos
    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='left', vertical='center')
    data_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definir encabezados
    headers = [
        'Código',
        'Nombre',
        'Categoría',
        'Proveedor',
        'Stock Actual',
        'Alerta de Stock',
        'Unidad de Medida',
        'Precio de Venta'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = header_style
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Escribir datos
    for row, producto in enumerate(productos, 2):
        cells = [
            (producto.código),
            (producto.nombre),
            (producto.categoría.nombre),
            (get_proveedor_nombre(producto.proveedor)),
            (producto.stock),
            (producto.alerta_stock),
            (producto.unidad_de_medida),
            (float(producto.precio_de_venta))
        ]
        
        for col, value in enumerate(cells, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.style = data_style
    
    # Ajustar el ancho de las columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Productos_Stock_Bajo.xlsx'
    
    # Guardar el libro de Excel
    wb.save(response)
    return response

@login_required
@never_cache
def exportar_productos_stock_bajo_categoria_excel(request, id):
    # Obtener la categoría y verificar que pertenece al usuario
    categoria = get_object_or_404(Categoria, id=id, user=request.user)
    
    # Obtener productos con stock bajo de la categoría
    productos = Producto.objects.filter(
        user=request.user,
        categoría=categoria,
        stock__lt=F('alerta_stock')
    ).select_related('categoría', 'proveedor').order_by('nombre')
    
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Stock Bajo - {categoria.nombre}"
    
    # Definir estilos
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_style.alignment = Alignment(horizontal='left', vertical='center')
    header_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Estilo para las celdas de datos
    data_style = NamedStyle(name='data_style')
    data_style.alignment = Alignment(horizontal='left', vertical='center')
    data_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definir encabezados
    headers = [
        'Código',
        'Nombre',
        'Proveedor',
        'Stock Actual',
        'Alerta de Stock',
        'Unidad de Medida',
        'Precio de Venta'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = header_style
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Escribir datos
    for row, producto in enumerate(productos, 2):
        cells = [
            (producto.código),
            (producto.nombre),
            (get_proveedor_nombre(producto.proveedor)),
            (producto.stock),
            (producto.alerta_stock),
            (producto.unidad_de_medida),
            (float(producto.precio_de_venta))
        ]
        
        for col, value in enumerate(cells, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.style = data_style
    
    # Ajustar el ancho de las columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Productos_Stock_Bajo_{categoria.nombre}.xlsx'
    
    # Guardar el libro de Excel
    wb.save(response)
    return response
